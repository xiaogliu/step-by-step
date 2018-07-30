# import openpyxl
# read model
# write model
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# read logic
wb_read = load_workbook('m_data.xlsx')
# wb_read = load_workbook('test.xlsx')
sheet_read = wb_read['Sheet1']

# write logic
wb_write = Workbook()
sheet_write = wb_write.active

# 统计学院
college_all = []
for i in sheet_read["F"]:
    college_all.append(i.value)
# 去重
colleges = list(set(college_all))
colleges.sort(key=college_all.index)
# print(colleges)

# 获取到行号就等于获取到了当前行，
# 缩小搜索范围可以统计好数据
# 科学学位/专业学位行号±
science_row_num = []
pro_row_num = []
for i in sheet_read["B"]:
    if i.value == '科学学位':
        science_row_num.append(i.row)
    else:
        pro_row_num.append(i.row)

# print(len(science_row_num) + len(pro_row_num))
# print(len(science_row_num))
# print(len(pro_row_num))

# 科学学位定向/非定向行号
science_dir_row_num = []
science_un_dir_row_num = []
for i in science_row_num:
    j = sheet_read.cell(row=i, column=column_index_from_string('L'))
    if j.value == '定向':
        science_dir_row_num.append(j.row)
    else:
        science_un_dir_row_num.append(j.row)

# 专业学位定向/非定向行号
pro_dir_row_num = []
pro_un_dir_row_num = []
for i in pro_row_num:
    j = sheet_read.cell(row=i, column=column_index_from_string('L'))
    if j.value == '定向':
        pro_dir_row_num.append(j.row)
    else:
        pro_un_dir_row_num.append(j.row)

# print(len(science_un_dir_row_num),len(science_dir_row_num))

# 科学学位定向统计
science_dir_tongkao_row_num = []
science_dir_dankao_row_num = []
science_dir_shaogan_row_num = []
for i in science_dir_row_num:
    j = sheet_read.cell(row=i, column=column_index_from_string('M'))
    if j.value == '':
        science_dir_tongkao_row_num.append(j.row)
    elif j.value == '单考':
        science_dir_dankao_row_num.append(j.row)
    elif j.value == '少干':
        science_dir_shaogan_row_num.append(j.row)

# 专业学位定向统计
pro_dir_tongkao_row_num = []
pro_dir_dankao_row_num = []
pro_dir_shaogan_row_num = []
pro_dir_tiaoji_row_num = []
for i in pro_dir_row_num:
    j = sheet_read.cell(row=i, column=column_index_from_string('M'))
    if j.value == '':
        pro_dir_tongkao_row_num.append(j.row)
    elif j.value == '单考':
        pro_dir_dankao_row_num.append(j.row)
    elif j.value == '少干':
        pro_dir_shaogan_row_num.append(j.row)
    elif j.value == '校内调剂':
        pro_dir_tiaoji_row_num.append(j.row)


# print(len(science_dir_row_num), len(science_dir_tongkao_row_num), len(science_dir_dankao_row_num), len(science_dir_shaogan_row_num))

def count(arr, c):
    # 科学学位定向单考统计
    middle = []
    for index, v in enumerate(colleges):
        middle.append(0)
    # del middle[0]
    # print(len(middle))
    for i in arr:
        j = sheet_read.cell(row=i, column=column_index_from_string('F'))
        for index, v in enumerate(colleges):
            if j.value == v:
                middle[index] += 1
    # 从 A4 开始，前面三行标题信息
    # print(middle)
    for i, v in enumerate(middle):
        if v == 0:
            middle[i] = ''
    for i, v in enumerate(middle):
        # print(i, v)
        # 从 A4 开始，前面三行标题信息
        sheet_write[(c + "%d") % (i + 4)].value = v


count(science_dir_tongkao_row_num, 'B')
count(science_dir_dankao_row_num, 'C')
count(science_dir_shaogan_row_num, 'D')

count(pro_dir_tongkao_row_num, 'K')
count(pro_dir_dankao_row_num, 'L')
count(pro_dir_shaogan_row_num, 'M')
count(pro_dir_tiaoji_row_num, 'N')

# 科学学位非定向统计
science_un_dir_tongkao_row_num = []
science_un_dir_tuimian_row_num = []
science_un_dir_tiaoji_row_num = []
for i in science_un_dir_row_num:
    j = sheet_read.cell(row=i, column=column_index_from_string('M'))
    if j.value == '':
        science_un_dir_tongkao_row_num.append(j.row)
    elif j.value == '推免':
        science_un_dir_tuimian_row_num.append(j.row)
    elif j.value == '校内调剂':
        science_un_dir_tiaoji_row_num.append(j.row)

# 专业学位非定向统计
pro_un_dir_tongkao_row_num = []
pro_un_dir_tuimian_row_num = []
pro_un_dir_tiaoji_row_num = []
pro_un_dir_shibing_row_num = []
for i in pro_un_dir_row_num:
    j = sheet_read.cell(row=i, column=column_index_from_string('M'))
    if j.value == '':
        pro_un_dir_tongkao_row_num.append(j.row)
    elif j.value == '推免':
        pro_un_dir_tuimian_row_num.append(j.row)
    elif j.value == '校内调剂':
        pro_un_dir_tiaoji_row_num.append(j.row)
    elif j.value == '士兵':
        pro_un_dir_shibing_row_num.append(j.row)


# print(len(science_un_dir_row_num), len(science_un_dir_tongkao_row_num), len(science_un_dir_tuimian_row_num), len(science_un_dir_tiaoji_row_num))

def count(arr, c):
    # 科学学位定向单考统计
    middle = []
    for index, v in enumerate(colleges):
        middle.append(0)
    # del middle[0]
    # print(len(middle))
    for i in arr:
        j = sheet_read.cell(row=i, column=column_index_from_string('F'))
        for index, v in enumerate(colleges):
            if j.value == v:
                middle[index] += 1
    # 从 A4 开始，前面三行标题信息
    # print(middle)
    for i, v in enumerate(middle):
        if v == 0:
            middle[i] = ''
    for i, v in enumerate(middle):
        # print(i, v)
        # 从 A4 开始，前面三行标题信息
        sheet_write[(c + "%d") % (i + 4)].value = v


count(science_un_dir_tongkao_row_num, 'F')
count(science_un_dir_tuimian_row_num, 'G')
count(science_un_dir_tiaoji_row_num, 'H')

count(pro_un_dir_tongkao_row_num, 'P')
count(pro_un_dir_tuimian_row_num, 'Q')
count(pro_un_dir_shibing_row_num, 'R')
count(pro_un_dir_tiaoji_row_num, 'S')

# 科学学位非定向备注
science_dir_mark_all = []
science_un_dir_mark_all = []
for i in science_dir_row_num:
    j = sheet_read.cell(row=i, column=column_index_from_string('M'))
    science_dir_mark_all.append(j.value)
# 去重
science_dir_mark = list(set(science_dir_mark_all))
science_dir_mark.sort(key=science_dir_mark_all.index)
# print(science_dir_mark)

# 科学学位非定向备注
science_un_dir_mark_all = []
for i in science_un_dir_row_num:
    j = sheet_read.cell(row=i, column=column_index_from_string('M'))
    science_un_dir_mark_all.append(j.value)
# 去重
science_un_dir_mark = list(set(science_un_dir_mark_all))
science_un_dir_mark.sort(key=science_un_dir_mark_all.index)
# print(science_un_dir_mark)

# 专业学位定向/非定向行号
pro_dir_row_num = []
pro_un_dir_row_num = []
for i in pro_row_num:
    j = sheet_read.cell(row=i, column=column_index_from_string('L'))
    if j.value == '定向':
        pro_dir_row_num.append(j.row)
    else:
        pro_un_dir_row_num.append(j.row)

# print(len(pro_un_dir_row_num),len(pro_dir_row_num))

# 科学学位定向备注
pro_dir_mark_all = []
pro_un_dir_mark_all = []
for i in pro_dir_row_num:
    j = sheet_read.cell(row=i, column=column_index_from_string('M'))
    pro_dir_mark_all.append(j.value)
# 去重
pro_dir_mark = list(set(pro_dir_mark_all))
pro_dir_mark.sort(key=pro_dir_mark_all.index)
# print(pro_dir_mark)

# 科学学位非定向备注
pro_un_dir_mark_all = []
for i in pro_un_dir_row_num:
    j = sheet_read.cell(row=i, column=column_index_from_string('M'))
    pro_un_dir_mark_all.append(j.value)
# 去重
pro_un_dir_mark = list(set(pro_un_dir_mark_all))
pro_un_dir_mark.sort(key=pro_un_dir_mark_all.index)
# print(pro_un_dir_mark)



# # write logic
# wb_write = Workbook()
# sheet_write = wb_write.active

sheet_write.title = '2018硕士拟录取统计'
sheet_write['A1'] = '2018硕士拟录取统计'

################################
# 科学学位
################################
sheet_write['B2'] = '科学学位'

sheet_write['B3'] = '定向'

# 科学学位定向定向类别
sheet_write['B4'] = '统考'
sheet_write['C4'] = '单考'
sheet_write['D4'] = '少干'
sheet_write['E4'] = '小计'

sheet_write['F3'] = '非定向'

sheet_write['F4'] = '统考'
sheet_write['G4'] = '推免'
sheet_write['H4'] = '校内调剂'
sheet_write['I4'] = '小计'
sheet_write['J3'] = '汇总'

################################
# 专业学位
################################
sheet_write['K2'] = '专业学位'

sheet_write['K3'] = '定向'

# 科学学位定向定向类别
sheet_write['K4'] = '统考'
sheet_write['L4'] = '单考'
sheet_write['M4'] = '少干'
sheet_write['N4'] = '校内调剂'
sheet_write['O4'] = '小计'

sheet_write['P3'] = '非定向'

sheet_write['P4'] = '统考'
sheet_write['Q4'] = '推免'
sheet_write['R4'] = '士兵'
sheet_write['S4'] = '校内调剂'
sheet_write['T4'] = '小计'
sheet_write['U3'] = '汇总'

# 添加学院
for i, v in enumerate(colleges):
    # print(i, v)
    # 从 A4 开始，前面三行标题信息
    sheet_write["A%d" % (i + 4)].value = v

# 保存 excel
wb_write.save('result.xlsx')
